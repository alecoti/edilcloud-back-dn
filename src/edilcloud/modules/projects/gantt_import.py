from __future__ import annotations

import csv
import json
import re
import subprocess
import sys
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from io import BytesIO, StringIO
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any
from xml.etree import ElementTree


SUPPORTED_GANTT_IMPORT_EXTENSIONS = {".csv", ".txt", ".xlsx", ".xls", ".xml", ".mpp"}
CSV_EXTENSIONS = {".csv", ".txt"}

HEADER_ALIASES = {
    "row_id": {"id", "uid", "unique id", "task id", "codice", "code"},
    "phase": {"fase", "phase", "summary", "summary task", "macrofase"},
    "activity": {
        "attivita",
        "attività",
        "task",
        "task name",
        "activity",
        "activity name",
        "name",
        "nome",
        "titolo",
    },
    "type": {"tipo", "type", "row type", "kind"},
    "level": {"livello", "level", "outline level", "indent level"},
    "summary": {"summary", "is summary", "phase row", "fase?"},
    "parent": {"parent", "parent id", "parent uid"},
    "company": {
        "azienda assegnataria",
        "azienda",
        "company",
        "company name",
        "assignee",
        "assegnatario",
    },
    "start": {
        "inizio",
        "data inizio",
        "start",
        "start date",
        "planned start",
        "date start",
    },
    "end": {
        "fine",
        "data fine",
        "finish",
        "end",
        "finish date",
        "planned finish",
        "date end",
    },
    "duration": {
        "durata",
        "durata (gg)",
        "duration",
        "duration (gg)",
        "duration days",
    },
    "progress": {
        "avanzamento",
        "percentuale avanzamento",
        "progress",
        "% complete",
        "percent complete",
    },
    "predecessors": {"predecessori", "predecessor", "predecessors", "depends on"},
    "link_type": {"tipo vincolo", "dependency type", "link type", "relation type"},
    "lag_days": {"lag", "lead/lag", "ritardo"},
    "description": {"description", "descrizione", "details", "dettagli"},
    "note": {"note", "nota", "notes", "commenti"},
}

PROJECT_LINK_TYPE_MAP = {
    "FS": "e2s",
    "SS": "s2s",
    "FF": "e2e",
    "SF": "s2e",
}
MS_PROJECT_LINK_TYPE_MAP = {
    0: "e2e",
    1: "e2s",
    2: "s2e",
    3: "s2s",
}
PERCENT_LAG_FORMATS = {19, 20, 35, 36, 51, 52}


@dataclass(slots=True)
class ImportWarning:
    code: str
    message: str
    level: str = "warning"


@dataclass(slots=True)
class ImportedLink:
    source_ref: str
    target_ref: str
    link_type: str = "e2s"
    lag_days: int = 0


@dataclass(slots=True)
class ImportedActivity:
    ref: str
    title: str
    date_start: date
    date_end: date
    progress: int = 0
    description: str = ""
    note: str = ""


@dataclass(slots=True)
class ImportedPhase:
    ref: str
    name: str
    date_start: date
    date_end: date
    progress: int = 0
    company_label: str = ""
    note: str = ""
    activities: list[ImportedActivity] = field(default_factory=list)


@dataclass(slots=True)
class ImportedPlan:
    detected_format: str
    source_system: str | None
    phases: list[ImportedPhase]
    links: list[ImportedLink]
    warnings: list[ImportWarning]
    detected_company_labels: list[str] = field(default_factory=list)


@dataclass(slots=True)
class _PendingLink:
    source_token: str
    target_ref: str
    link_type: str
    lag_days: int


def parse_gantt_import_file(uploaded_file) -> ImportedPlan:
    file_name = getattr(uploaded_file, "name", "") or "gantt-import"
    extension = Path(file_name).suffix.lower()
    if extension not in SUPPORTED_GANTT_IMPORT_EXTENSIONS:
        raise ValueError(
            "Formato non supportato. Usa un file CSV, XLS, XLSX, XML o MPP Microsoft Project."
        )

    file_bytes = uploaded_file.read()
    if not file_bytes:
        raise ValueError("Il file selezionato e vuoto.")

    if extension in CSV_EXTENSIONS:
        return _parse_csv_import(file_bytes)
    if extension == ".xlsx":
        return _parse_xlsx_import(file_bytes)
    if extension == ".xls":
        return _parse_xls_import(file_bytes)
    if extension == ".mpp":
        return _parse_ms_project_mpp_import(file_bytes)
    return _parse_ms_project_xml_import(file_bytes)


def _parse_csv_import(file_bytes: bytes) -> ImportedPlan:
    decoded = _decode_text(file_bytes)
    try:
        dialect = csv.Sniffer().sniff(decoded[:4096], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";" if ";" in decoded[:2048] else ","
    rows = list(csv.reader(StringIO(decoded), dialect))
    return _parse_tabular_import(rows, detected_format="csv", source_system="spreadsheet")


def _parse_xlsx_import(file_bytes: bytes) -> ImportedPlan:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Supporto XLSX non disponibile sul backend.") from exc

    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    worksheet = _first_non_empty_openpyxl_sheet(workbook)
    if worksheet is None:
        raise ValueError("Il file XLSX non contiene fogli valorizzati.")
    rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    return _parse_tabular_import(rows, detected_format="xlsx", source_system="spreadsheet")


def _parse_xls_import(file_bytes: bytes) -> ImportedPlan:
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError("Supporto XLS non disponibile sul backend.") from exc

    workbook = xlrd.open_workbook(file_contents=file_bytes)
    sheet = _first_non_empty_xlrd_sheet(workbook)
    if sheet is None:
        raise ValueError("Il file XLS non contiene fogli valorizzati.")

    rows: list[list[Any]] = []
    for row_index in range(sheet.nrows):
        row_values: list[Any] = []
        for column_index in range(sheet.ncols):
            cell = sheet.cell(row_index, column_index)
            if cell.ctype == xlrd.XL_CELL_DATE:
                row_values.append(xlrd.xldate_as_datetime(cell.value, workbook.datemode))
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                numeric = float(cell.value)
                row_values.append(int(numeric) if numeric.is_integer() else numeric)
            elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                row_values.append(bool(cell.value))
            else:
                row_values.append(cell.value)
        rows.append(row_values)

    return _parse_tabular_import(rows, detected_format="xls", source_system="spreadsheet")


def _parse_ms_project_xml_import(file_bytes: bytes) -> ImportedPlan:
    decoded = _decode_text(file_bytes)
    root = ElementTree.fromstring(decoded)
    namespace = _xml_namespace(root.tag)
    ns = {"m": namespace} if namespace else {}

    def find_text(node: ElementTree.Element, tag_name: str) -> str:
        selector = f"m:{tag_name}" if namespace else tag_name
        child = node.find(selector, ns)
        return _clean_text(child.text if child is not None else "")

    minutes_per_day = _parse_int(find_text(root, "MinutesPerDay")) or 480
    raw_tasks: list[dict[str, Any]] = []
    task_selector = ".//m:Tasks/m:Task" if namespace else ".//Tasks/Task"
    for task_element in root.findall(task_selector, ns):
        uid = find_text(task_element, "UID")
        name = find_text(task_element, "Name")
        if not uid or uid == "0" or not name:
            continue

        predecessors: list[tuple[str, str, int]] = []
        predecessor_selector = "m:PredecessorLink" if namespace else "PredecessorLink"
        for predecessor_link in task_element.findall(predecessor_selector, ns):
            predecessor_uid = find_text(predecessor_link, "PredecessorUID")
            if not predecessor_uid:
                continue
            raw_type = _parse_int(find_text(predecessor_link, "Type"))
            lag_raw = _parse_int(find_text(predecessor_link, "LinkLag")) or 0
            lag_format = _parse_int(find_text(predecessor_link, "LagFormat"))
            predecessors.append(
                (
                    predecessor_uid,
                    MS_PROJECT_LINK_TYPE_MAP.get(raw_type or 1, "e2s"),
                    _msp_link_lag_to_days(lag_raw, lag_format, minutes_per_day),
                )
            )

        raw_tasks.append(
            {
                "uid": uid,
                "name": name,
                "summary": find_text(task_element, "Summary") == "1",
                "outline_level": _parse_int(find_text(task_element, "OutlineLevel")) or 1,
                "start": _parse_date(find_text(task_element, "Start")),
                "end": _parse_date(find_text(task_element, "Finish")),
                "progress": _clamp_progress(_parse_float(find_text(task_element, "PercentComplete")) or 0),
                "note": find_text(task_element, "Notes"),
                "company": (
                    find_text(task_element, "Text1")
                    or find_text(task_element, "Text2")
                    or find_text(task_element, "Text3")
                ),
                "predecessors": predecessors,
            }
        )

    if not raw_tasks:
        raise ValueError("Nessun task Microsoft Project trovato nel file XML.")

    return _build_microsoft_project_plan(
        raw_tasks,
        minutes_per_day=minutes_per_day,
        detected_format="ms-project-xml",
    )


def _parse_ms_project_mpp_import(file_bytes: bytes) -> ImportedPlan:
    raw_tasks, minutes_per_day = _read_mpp_raw_tasks_with_worker(file_bytes)
    if not raw_tasks:
        raise ValueError("Nessun task Microsoft Project trovato nel file MPP.")
    return _build_microsoft_project_plan(
        raw_tasks,
        minutes_per_day=minutes_per_day,
        detected_format="ms-project-mpp",
    )


def _read_mpp_raw_tasks_with_worker(file_bytes: bytes) -> tuple[list[dict[str, Any]], int]:
    with NamedTemporaryFile(delete=False, suffix=".mpp") as temp_file:
        temp_file.write(file_bytes)
        temp_path = Path(temp_file.name)

    try:
        worker_path = Path(__file__).with_name("mpp_reader_worker.py")
        command = [
            sys.executable,
            str(worker_path),
            str(temp_path),
        ]
        try:
            result = subprocess.run(
                command,
                capture_output=True,
                text=True,
                timeout=45,
                check=False,
            )
        except subprocess.TimeoutExpired as exc:
            raise ValueError("Lettura file MPP scaduta. Riprova con un file piu piccolo.") from exc
    finally:
        try:
            temp_path.unlink(missing_ok=True)
        except OSError:
            pass

    if result.returncode != 0:
        detail = (result.stderr or "").strip() or "Impossibile leggere il file MPP."
        raise ValueError(detail)

    output = (result.stdout or "").strip()
    if output and not output.lstrip().startswith("{"):
        for line in reversed(output.splitlines()):
            candidate = line.strip()
            if candidate.startswith("{"):
                output = candidate
                break

    try:
        payload = json.loads(output or "{}")
    except json.JSONDecodeError as exc:
        raise ValueError("Risposta parser MPP non valida.") from exc

    raw_tasks = payload.get("raw_tasks")
    if not isinstance(raw_tasks, list):
        raise ValueError(
            "Impossibile leggere il file MPP. Verifica che non sia corrotto o protetto da password."
        )

    minutes_per_day = _parse_int(payload.get("minutes_per_day")) or 480
    return raw_tasks, minutes_per_day


def _parse_tabular_import(
    raw_rows: list[list[Any]],
    *,
    detected_format: str,
    source_system: str,
) -> ImportedPlan:
    rows = [_normalize_row(row) for row in raw_rows if any(_clean_text(cell) for cell in row)]
    if len(rows) < 2:
        raise ValueError("Il file non contiene abbastanza righe da importare.")

    headers = [_normalize_header(cell) for cell in rows[0]]
    column_map = _build_column_map(headers)
    body_rows = rows[1:]

    if column_map.get("phase") is not None and column_map.get("activity") is not None:
        return _parse_phase_activity_rows(
            body_rows,
            column_map=column_map,
            detected_format=detected_format,
            source_system=source_system,
        )

    if any(column_map.get(key) is not None for key in ("level", "summary", "type", "parent")):
        return _parse_tree_rows(
            body_rows,
            column_map=column_map,
            detected_format=detected_format,
            source_system=source_system,
        )

    if column_map.get("activity") is None:
        raise ValueError(
            "Non trovo una colonna compatibile con Nome task/Attivita nel file caricato."
        )

    return _parse_flat_rows(
        body_rows,
        column_map=column_map,
        detected_format=detected_format,
        source_system=source_system,
    )


def _build_microsoft_project_plan(
    raw_tasks: list[dict[str, Any]],
    *,
    minutes_per_day: int,
    detected_format: str,
) -> ImportedPlan:
    del minutes_per_day
    phases: list[ImportedPhase] = []
    links: list[ImportedLink] = []
    warnings: list[ImportWarning] = []
    company_labels: list[str] = []
    counters = {"phase": 0, "activity": 0}
    uid_to_ref: dict[str, str] = {}
    name_to_ref: dict[str, str] = {}
    summary_stack: list[tuple[int, ImportedPhase]] = []
    pending_links: list[_PendingLink] = []

    def next_ref(kind: str) -> str:
        counters[kind] += 1
        return f"{kind}:{counters[kind]}"

    def register_reference(token: str, ref: str, name: str) -> None:
        normalized_token = _normalize_lookup(token)
        if normalized_token:
            uid_to_ref[normalized_token] = ref
        normalized_name = _normalize_lookup(name)
        if normalized_name and normalized_name not in name_to_ref:
            name_to_ref[normalized_name] = ref

    for raw_task in raw_tasks:
        level = raw_task["outline_level"]
        while summary_stack and summary_stack[-1][0] >= level:
            summary_stack.pop()

        start, end = _resolve_dates(
            _parse_date(raw_task["start"]),
            _parse_date(raw_task["end"]),
            None,
        )
        company_label = _clean_text(raw_task["company"])
        if company_label and company_label not in company_labels:
            company_labels.append(company_label)

        if raw_task["summary"]:
            if not summary_stack:
                phase = ImportedPhase(
                    ref=next_ref("phase"),
                    name=raw_task["name"],
                    date_start=start,
                    date_end=end,
                    progress=raw_task["progress"],
                    company_label=company_label,
                    note=raw_task["note"],
                )
                phases.append(phase)
                summary_stack.append((level, phase))
                register_reference(raw_task["uid"], phase.ref, phase.name)
            else:
                top_phase = summary_stack[0][1]
                register_reference(raw_task["uid"], top_phase.ref, raw_task["name"])
                summary_stack.append((level, top_phase))
            continue

        if summary_stack:
            phase = summary_stack[0][1]
            activity = ImportedActivity(
                ref=next_ref("activity"),
                title=raw_task["name"],
                date_start=start,
                date_end=end,
                progress=raw_task["progress"],
                note=raw_task["note"],
            )
            phase.activities.append(activity)
            phase.date_start = min(phase.date_start, start)
            phase.date_end = max(phase.date_end, end)
            register_reference(raw_task["uid"], activity.ref, activity.title)
            target_ref = activity.ref
        else:
            phase = ImportedPhase(
                ref=next_ref("phase"),
                name=raw_task["name"],
                date_start=start,
                date_end=end,
                progress=raw_task["progress"],
                company_label=company_label,
                note=raw_task["note"],
            )
            phases.append(phase)
            register_reference(raw_task["uid"], phase.ref, phase.name)
            target_ref = phase.ref

        for predecessor_uid, link_type, lag_days in raw_task["predecessors"]:
            pending_links.append(
                _PendingLink(
                    source_token=predecessor_uid,
                    target_ref=target_ref,
                    link_type=link_type,
                    lag_days=lag_days,
                )
            )

    _normalize_phase_ranges(phases)
    links.extend(_resolve_pending_links(pending_links, uid_to_ref, name_to_ref, warnings))
    return ImportedPlan(
        detected_format=detected_format,
        source_system="microsoft-project",
        phases=phases,
        links=links,
        warnings=warnings,
        detected_company_labels=company_labels,
    )


def _parse_phase_activity_rows(
    rows: list[list[Any]],
    *,
    column_map: dict[str, int],
    detected_format: str,
    source_system: str,
) -> ImportedPlan:
    phases: list[ImportedPhase] = []
    warnings: list[ImportWarning] = []
    links: list[ImportedLink] = []
    company_labels: list[str] = []
    phase_lookup: dict[str, ImportedPhase] = {}
    id_to_ref: dict[str, str] = {}
    name_to_ref: dict[str, str] = {}
    pending_links: list[_PendingLink] = []
    current_phase_key = ""
    counters = {"phase": 0, "activity": 0}

    def next_ref(kind: str) -> str:
        counters[kind] += 1
        return f"{kind}:{counters[kind]}"

    def register_reference(token: str, ref: str, name: str) -> None:
        normalized_token = _normalize_lookup(token)
        if normalized_token:
            id_to_ref[normalized_token] = ref
        normalized_name = _normalize_lookup(name)
        if normalized_name and normalized_name not in name_to_ref:
            name_to_ref[normalized_name] = ref

    def ensure_phase(phase_name: str, company_label: str, row_id: str) -> ImportedPhase:
        nonlocal current_phase_key
        normalized_key = _normalize_lookup(phase_name) or _normalize_lookup(row_id or current_phase_key)
        if not normalized_key:
            normalized_key = f"phase-key-{len(phases) + 1}"
        current_phase_key = normalized_key
        existing = phase_lookup.get(normalized_key)
        if existing is not None:
            if company_label and not existing.company_label:
                existing.company_label = company_label
            return existing

        phase = ImportedPhase(
            ref=next_ref("phase"),
            name=phase_name or f"Fase {len(phases) + 1}",
            date_start=date.today(),
            date_end=date.today(),
            company_label=company_label,
        )
        phases.append(phase)
        phase_lookup[normalized_key] = phase
        register_reference(row_id or phase_name, phase.ref, phase.name)
        return phase

    for raw_row in rows:
        phase_name = _cell(raw_row, column_map, "phase")
        activity_name = _cell(raw_row, column_map, "activity")
        row_id = _cell(raw_row, column_map, "row_id")
        company_label = _cell(raw_row, column_map, "company")
        note = _cell(raw_row, column_map, "note")
        description = _cell(raw_row, column_map, "description")
        if company_label and company_label not in company_labels:
            company_labels.append(company_label)

        if phase_name:
            phase = ensure_phase(phase_name, company_label, row_id)
        elif current_phase_key:
            phase = phase_lookup[current_phase_key]
        else:
            warnings.append(
                ImportWarning(
                    code="missing-phase",
                    message="Una riga attivita e stata ignorata perche non ha una fase di riferimento.",
                )
            )
            continue

        start = _parse_date(_cell(raw_row, column_map, "start"))
        end = _parse_date(_cell(raw_row, column_map, "end"))
        duration_days = _parse_duration_days(_cell(raw_row, column_map, "duration"))
        progress = _clamp_progress(_parse_float(_cell(raw_row, column_map, "progress")) or 0)

        if not activity_name:
            phase_start, phase_end = _resolve_dates(start, end, duration_days)
            phase.date_start = min(phase.date_start, phase_start)
            phase.date_end = max(phase.date_end, phase_end)
            if note and not phase.note:
                phase.note = note
            register_reference(row_id or phase.name, phase.ref, phase.name)
            predecessor_value = _cell(raw_row, column_map, "predecessors")
            row_link_type = _cell(raw_row, column_map, "link_type")
            row_lag_days = _parse_duration_days(_cell(raw_row, column_map, "lag_days")) or 0
            for source_token, link_type, lag_days in _parse_predecessor_tokens(
                predecessor_value,
                default_type=row_link_type,
                default_lag=row_lag_days,
            ):
                pending_links.append(
                    _PendingLink(
                        source_token=source_token,
                        target_ref=phase.ref,
                        link_type=link_type,
                        lag_days=lag_days,
                    )
                )
            continue

        activity_start, activity_end = _resolve_dates(start, end, duration_days)
        activity = ImportedActivity(
            ref=next_ref("activity"),
            title=activity_name,
            date_start=activity_start,
            date_end=activity_end,
            progress=progress,
            description=description,
            note=note,
        )
        phase.activities.append(activity)
        phase.date_start = min(phase.date_start, activity_start)
        phase.date_end = max(phase.date_end, activity_end)
        register_reference(row_id or activity_name, activity.ref, activity.title)

        predecessor_value = _cell(raw_row, column_map, "predecessors")
        row_link_type = _cell(raw_row, column_map, "link_type")
        row_lag_days = _parse_duration_days(_cell(raw_row, column_map, "lag_days")) or 0
        for source_token, link_type, lag_days in _parse_predecessor_tokens(
            predecessor_value,
            default_type=row_link_type,
            default_lag=row_lag_days,
        ):
            pending_links.append(
                _PendingLink(
                    source_token=source_token,
                    target_ref=activity.ref,
                    link_type=link_type,
                    lag_days=lag_days,
                )
            )

    _normalize_phase_ranges(phases)
    links.extend(_resolve_pending_links(pending_links, id_to_ref, name_to_ref, warnings))
    return ImportedPlan(
        detected_format=detected_format,
        source_system=source_system,
        phases=phases,
        links=links,
        warnings=warnings,
        detected_company_labels=company_labels,
    )


def _parse_tree_rows(
    rows: list[list[Any]],
    *,
    column_map: dict[str, int],
    detected_format: str,
    source_system: str,
) -> ImportedPlan:
    phases: list[ImportedPhase] = []
    warnings: list[ImportWarning] = []
    links: list[ImportedLink] = []
    company_labels: list[str] = []
    pending_links: list[_PendingLink] = []
    id_to_ref: dict[str, str] = {}
    name_to_ref: dict[str, str] = {}
    counters = {"phase": 0, "activity": 0}
    current_phase: ImportedPhase | None = None

    def next_ref(kind: str) -> str:
        counters[kind] += 1
        return f"{kind}:{counters[kind]}"

    def register_reference(token: str, ref: str, name: str) -> None:
        normalized_token = _normalize_lookup(token)
        if normalized_token:
            id_to_ref[normalized_token] = ref
        normalized_name = _normalize_lookup(name)
        if normalized_name and normalized_name not in name_to_ref:
            name_to_ref[normalized_name] = ref

    for raw_row in rows:
        name = _cell(raw_row, column_map, "activity")
        if not name:
            continue

        row_id = _cell(raw_row, column_map, "row_id")
        company_label = _cell(raw_row, column_map, "company")
        if company_label and company_label not in company_labels:
            company_labels.append(company_label)

        level = _parse_int(_cell(raw_row, column_map, "level")) or 1
        summary_flag = _parse_bool(_cell(raw_row, column_map, "summary"))
        row_type = _normalize_lookup(_cell(raw_row, column_map, "type"))
        is_summary = summary_flag or row_type in {"phase", "summary", "group", "fase"} or level <= 1

        start = _parse_date(_cell(raw_row, column_map, "start"))
        end = _parse_date(_cell(raw_row, column_map, "end"))
        duration_days = _parse_duration_days(_cell(raw_row, column_map, "duration"))
        progress = _clamp_progress(_parse_float(_cell(raw_row, column_map, "progress")) or 0)
        note = _cell(raw_row, column_map, "note")
        description = _cell(raw_row, column_map, "description")
        resolved_start, resolved_end = _resolve_dates(start, end, duration_days)

        if is_summary or current_phase is None:
            current_phase = ImportedPhase(
                ref=next_ref("phase"),
                name=name,
                date_start=resolved_start,
                date_end=resolved_end,
                progress=progress,
                company_label=company_label,
                note=note,
            )
            phases.append(current_phase)
            register_reference(row_id or name, current_phase.ref, current_phase.name)
            target_ref = current_phase.ref
        else:
            activity = ImportedActivity(
                ref=next_ref("activity"),
                title=name,
                date_start=resolved_start,
                date_end=resolved_end,
                progress=progress,
                description=description,
                note=note,
            )
            current_phase.activities.append(activity)
            current_phase.date_start = min(current_phase.date_start, resolved_start)
            current_phase.date_end = max(current_phase.date_end, resolved_end)
            register_reference(row_id or name, activity.ref, activity.title)
            target_ref = activity.ref

        predecessor_value = _cell(raw_row, column_map, "predecessors")
        row_link_type = _cell(raw_row, column_map, "link_type")
        row_lag_days = _parse_duration_days(_cell(raw_row, column_map, "lag_days")) or 0
        for source_token, link_type, lag_days in _parse_predecessor_tokens(
            predecessor_value,
            default_type=row_link_type,
            default_lag=row_lag_days,
        ):
            pending_links.append(
                _PendingLink(
                    source_token=source_token,
                    target_ref=target_ref,
                    link_type=link_type,
                    lag_days=lag_days,
                )
            )

    _normalize_phase_ranges(phases)
    links.extend(_resolve_pending_links(pending_links, id_to_ref, name_to_ref, warnings))
    return ImportedPlan(
        detected_format=detected_format,
        source_system=source_system,
        phases=phases,
        links=links,
        warnings=warnings,
        detected_company_labels=company_labels,
    )


def _parse_flat_rows(
    rows: list[list[Any]],
    *,
    column_map: dict[str, int],
    detected_format: str,
    source_system: str,
) -> ImportedPlan:
    phases: list[ImportedPhase] = []
    warnings: list[ImportWarning] = []
    links: list[ImportedLink] = []
    company_labels: list[str] = []
    counters = {"phase": 0}
    pending_links: list[_PendingLink] = []
    id_to_ref: dict[str, str] = {}
    name_to_ref: dict[str, str] = {}

    def next_phase_ref() -> str:
        counters["phase"] += 1
        return f"phase:{counters['phase']}"

    for raw_row in rows:
        name = _cell(raw_row, column_map, "activity")
        if not name:
            continue

        row_id = _cell(raw_row, column_map, "row_id")
        company_label = _cell(raw_row, column_map, "company")
        if company_label and company_label not in company_labels:
            company_labels.append(company_label)

        start = _parse_date(_cell(raw_row, column_map, "start"))
        end = _parse_date(_cell(raw_row, column_map, "end"))
        duration_days = _parse_duration_days(_cell(raw_row, column_map, "duration"))
        progress = _clamp_progress(_parse_float(_cell(raw_row, column_map, "progress")) or 0)
        note = _cell(raw_row, column_map, "note")
        resolved_start, resolved_end = _resolve_dates(start, end, duration_days)

        phase = ImportedPhase(
            ref=next_phase_ref(),
            name=name,
            date_start=resolved_start,
            date_end=resolved_end,
            progress=progress,
            company_label=company_label,
            note=note,
        )
        phases.append(phase)

        normalized_row_id = _normalize_lookup(row_id)
        if normalized_row_id:
            id_to_ref[normalized_row_id] = phase.ref
        normalized_name = _normalize_lookup(name)
        if normalized_name and normalized_name not in name_to_ref:
            name_to_ref[normalized_name] = phase.ref

        predecessor_value = _cell(raw_row, column_map, "predecessors")
        row_link_type = _cell(raw_row, column_map, "link_type")
        row_lag_days = _parse_duration_days(_cell(raw_row, column_map, "lag_days")) or 0
        for source_token, link_type, lag_days in _parse_predecessor_tokens(
            predecessor_value,
            default_type=row_link_type,
            default_lag=row_lag_days,
        ):
            pending_links.append(
                _PendingLink(
                    source_token=source_token,
                    target_ref=phase.ref,
                    link_type=link_type,
                    lag_days=lag_days,
                )
            )

    links.extend(_resolve_pending_links(pending_links, id_to_ref, name_to_ref, warnings))
    return ImportedPlan(
        detected_format=detected_format,
        source_system=source_system,
        phases=phases,
        links=links,
        warnings=warnings,
        detected_company_labels=company_labels,
    )


def _resolve_pending_links(
    pending_links: list[_PendingLink],
    id_to_ref: dict[str, str],
    name_to_ref: dict[str, str],
    warnings: list[ImportWarning],
) -> list[ImportedLink]:
    resolved_links: list[ImportedLink] = []
    seen_pairs: set[tuple[str, str, str, int]] = set()

    for pending_link in pending_links:
        normalized_source = _normalize_lookup(pending_link.source_token)
        source_ref = id_to_ref.get(normalized_source) or name_to_ref.get(normalized_source)
        if not source_ref:
            warnings.append(
                ImportWarning(
                    code="unresolved-predecessor",
                    message=f"Vincolo ignorato: predecessore '{pending_link.source_token}' non trovato.",
                )
            )
            continue
        if source_ref == pending_link.target_ref:
            continue
        key = (source_ref, pending_link.target_ref, pending_link.link_type, pending_link.lag_days)
        if key in seen_pairs:
            continue
        seen_pairs.add(key)
        resolved_links.append(
            ImportedLink(
                source_ref=source_ref,
                target_ref=pending_link.target_ref,
                link_type=pending_link.link_type,
                lag_days=pending_link.lag_days,
            )
        )

    return resolved_links


def _parse_predecessor_tokens(
    raw_value: str,
    *,
    default_type: str,
    default_lag: int,
) -> list[tuple[str, str, int]]:
    normalized = _clean_text(raw_value)
    if not normalized:
        return []

    link_type = PROJECT_LINK_TYPE_MAP.get(_clean_text(default_type).upper(), "e2s")
    items: list[tuple[str, str, int]] = []
    for token in re.split(r"[;,]", normalized):
        candidate = _clean_text(token)
        if not candidate:
            continue
        match = re.match(
            r"^(?P<ref>.+?)(?P<type>FS|SS|FF|SF)?(?P<lag>[+-]\s*\d+(?:[.,]\d+)?)?$",
            candidate,
            flags=re.IGNORECASE,
        )
        if not match:
            items.append((candidate, link_type, default_lag))
            continue
        resolved_type = PROJECT_LINK_TYPE_MAP.get(
            _clean_text(match.group("type")).upper(),
            link_type,
        )
        resolved_lag = default_lag
        if match.group("lag"):
            resolved_lag = int(round(float(match.group("lag").replace(" ", "").replace(",", "."))))
        items.append((match.group("ref"), resolved_type, resolved_lag))
    return items


def _build_column_map(headers: list[str]) -> dict[str, int]:
    result: dict[str, int] = {}
    for column_name, aliases in HEADER_ALIASES.items():
        for index, header in enumerate(headers):
            if header in aliases:
                result[column_name] = index
                break
    return result


def _normalize_row(row: list[Any]) -> list[Any]:
    return [cell for cell in row]


def _normalize_header(value: Any) -> str:
    return _normalize_lookup(value)


def _cell(row: list[Any], column_map: dict[str, int], key: str) -> str:
    index = column_map.get(key)
    if index is None or index >= len(row):
        return ""
    return _clean_text(row[index])


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    return str(value).strip()


def _normalize_lookup(value: Any) -> str:
    return _clean_text(value).lower().replace("_", " ").replace("-", " ")


def _parse_bool(value: str) -> bool:
    return _normalize_lookup(value) in {"1", "true", "yes", "si", "sì", "x"}


def _parse_int(value: str) -> int | None:
    try:
        return int(float(_clean_text(value)))
    except (TypeError, ValueError):
        return None


def _parse_float(value: str) -> float | None:
    normalized = _clean_text(value).replace("%", "").replace(",", ".")
    if not normalized:
        return None
    try:
        return float(normalized)
    except ValueError:
        return None


def _parse_duration_days(value: str) -> int | None:
    normalized = _clean_text(value)
    if not normalized:
        return None
    match = re.search(r"-?\d+(?:[.,]\d+)?", normalized)
    if not match:
        return None
    return max(0, int(round(float(match.group(0).replace(",", ".")))))


def _parse_date(value: Any) -> date | None:
    if value in {None, ""}:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    normalized = _clean_text(value).replace("Z", "+00:00")
    if not normalized:
        return None

    if "T" in normalized or " " in normalized:
        try:
            return datetime.fromisoformat(normalized).date()
        except ValueError:
            pass

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(normalized, fmt).date()
        except ValueError:
            continue
    return None


def _resolve_dates(
    start: date | None,
    end: date | None,
    duration_days: int | None,
) -> tuple[date, date]:
    baseline = date.today()
    resolved_start = start or end or baseline
    resolved_end = end or start or baseline
    if duration_days and duration_days > 0:
        if start and not end:
            resolved_end = resolved_start + timedelta(days=duration_days - 1)
        elif end and not start:
            resolved_start = resolved_end - timedelta(days=duration_days - 1)
    if resolved_end < resolved_start:
        return resolved_start, resolved_start
    return resolved_start, resolved_end


def _clamp_progress(value: float | int | None) -> int:
    try:
        numeric = int(round(float(value or 0)))
    except (TypeError, ValueError):
        numeric = 0
    return max(0, min(100, numeric))


def _normalize_phase_ranges(phases: list[ImportedPhase]) -> None:
    for phase in phases:
        if phase.activities:
            phase.date_start = min(activity.date_start for activity in phase.activities)
            phase.date_end = max(activity.date_end for activity in phase.activities)
        if phase.date_end < phase.date_start:
            phase.date_end = phase.date_start


def _decode_text(file_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-16", "cp1252", "latin-1"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("Impossibile leggere il file testuale caricato.")


def _first_non_empty_openpyxl_sheet(workbook):
    for worksheet in workbook.worksheets:
        if worksheet.max_row and worksheet.max_column:
            return worksheet
    return None


def _first_non_empty_xlrd_sheet(workbook):
    for worksheet in workbook.sheets():
        if worksheet.nrows and worksheet.ncols:
            return worksheet
    return None


def _xml_namespace(tag_name: str) -> str:
    if tag_name.startswith("{") and "}" in tag_name:
        return tag_name[1 : tag_name.index("}")]
    return ""


def _msp_link_lag_to_days(link_lag: int, lag_format: int | None, minutes_per_day: int) -> int:
    if link_lag == 0:
        return 0
    if lag_format in PERCENT_LAG_FORMATS:
        return 0
    denominator = max(minutes_per_day * 10, 1)
    value = link_lag / denominator
    if 0 < abs(value) < 0.5:
        return 1 if value > 0 else -1
    return int(round(value))
